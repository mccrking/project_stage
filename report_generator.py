from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

class ReportGenerator:
    def __init__(self):
        self.reports_dir = 'reports'
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_pdf_report(self, data, filename):
        """
        Génère un rapport PDF
        
        Args:
            data (dict): Données du rapport
            filename (str): Nom du fichier PDF
        """
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # En-tête
            self._add_pdf_header(pdf, data)
            
            # Statistiques
            self._add_pdf_statistics(pdf, data)
            
            # Liste des appareils
            self._add_pdf_devices_table(pdf, data)
            
            # Pied de page
            self._add_pdf_footer(pdf)
            
            # Sauvegarder le fichier
            filepath = os.path.join(self.reports_dir, filename)
            pdf.output(filepath)
            
            print(f"📄 Rapport PDF généré: {filepath}")
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du rapport PDF: {str(e)}")
    
    def generate_excel_report(self, data, filename):
        """
        Génère un rapport Excel
        
        Args:
            data (dict): Données du rapport
            filename (str): Nom du fichier Excel
        """
        try:
            wb = openpyxl.Workbook()
            
            # Feuille de statistiques
            self._add_excel_statistics_sheet(wb, data)
            
            # Feuille des appareils
            self._add_excel_devices_sheet(wb, data)
            
            # Feuille de résumé
            self._add_excel_summary_sheet(wb, data)
            
            # Sauvegarder le fichier
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            print(f"📊 Rapport Excel généré: {filepath}")
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du rapport Excel: {str(e)}")
    
    def _add_pdf_header(self, pdf, data):
        """Ajoute l'en-tête au rapport PDF"""
        # Titre principal
        pdf.set_font('Arial', 'B', 20)
        pdf.cell(0, 20, 'CENTRAL DANONE - RAPPORT DE SUPERVISION RÉSEAU', ln=True, align='C')
        
        # Date du rapport
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, f'Date du rapport: {data["date"]}', ln=True, align='C')
        pdf.cell(0, 10, f'Généré le: {datetime.now().strftime("%d/%m/%Y à %H:%M")}', ln=True, align='C')
        
        pdf.ln(10)
    
    def _add_pdf_statistics(self, pdf, data):
        """Ajoute les statistiques au rapport PDF"""
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'STATISTIQUES GÉNÉRALES', ln=True)
        
        pdf.set_font('Arial', '', 12)
        stats = data['statistics']
        
        # Tableau des statistiques
        pdf.cell(60, 8, 'Total des appareils:', border=1)
        pdf.cell(30, 8, str(stats['total_devices']), border=1, ln=True)
        
        pdf.cell(60, 8, 'Appareils en ligne:', border=1)
        pdf.cell(30, 8, str(stats['online_devices']), border=1, ln=True)
        
        pdf.cell(60, 8, 'Appareils hors ligne:', border=1)
        pdf.cell(30, 8, str(stats['offline_devices']), border=1, ln=True)
        
        if 'scans_today' in stats:
            pdf.cell(60, 8, 'Scans effectués aujourd\'hui:', border=1)
            pdf.cell(30, 8, str(stats['scans_today']), border=1, ln=True)
        
        # Calcul du pourcentage de disponibilité
        if stats['total_devices'] > 0:
            uptime_percentage = (stats['online_devices'] / stats['total_devices']) * 100
            pdf.cell(60, 8, 'Taux de disponibilité:', border=1)
            pdf.cell(30, 8, f"{uptime_percentage:.1f}%", border=1, ln=True)
        
        pdf.ln(10)
    
    def _add_pdf_devices_table(self, pdf, data):
        """Ajoute le tableau des appareils au rapport PDF"""
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'LISTE DES APPAREILS', ln=True)
        
        # En-têtes du tableau
        pdf.set_font('Arial', 'B', 10)
        headers = ['Adresse IP', 'Nom d\'hôte', 'Statut', 'Dernière vue']
        widths = [40, 50, 30, 40]
        
        for i, header in enumerate(headers):
            pdf.cell(widths[i], 8, header, border=1)
        pdf.ln()
        
        # Données des appareils
        pdf.set_font('Arial', '', 9)
        for device in data['devices']:
            # Couleur selon le statut
            if device['status'] == 'online':
                pdf.set_fill_color(200, 255, 200)  # Vert clair
            else:
                pdf.set_fill_color(255, 200, 200)  # Rouge clair
            
            pdf.cell(widths[0], 6, device['ip'], border=1, fill=True)
            pdf.cell(widths[1], 6, device['hostname'][:25], border=1, fill=True)
            pdf.cell(widths[2], 6, device['status'], border=1, fill=True)
            pdf.cell(widths[3], 6, device['last_seen'], border=1, fill=True)
            pdf.ln()
        
        pdf.ln(10)
    
    def _add_pdf_footer(self, pdf):
        """Ajoute le pied de page au rapport PDF"""
        pdf.set_y(-30)
        pdf.set_font('Arial', 'I', 8)
        pdf.cell(0, 10, 'Rapport généré automatiquement par le système de supervision Central Danone', ln=True, align='C')
        pdf.cell(0, 10, 'Pour toute question, contactez l\'équipe IT', ln=True, align='C')
    
    def _add_excel_statistics_sheet(self, wb, data):
        """Ajoute la feuille de statistiques au rapport Excel"""
        ws = wb.active
        ws.title = "Statistiques"
        
        # Titre
        ws['A1'] = 'CENTRAL DANONE - RAPPORT DE SUPERVISION RÉSEAU'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A3'] = f'Date du rapport: {data["date"]}'
        ws['A4'] = f'Généré le: {datetime.now().strftime("%d/%m/%Y à %H:%M")}'
        
        # Statistiques
        ws['A6'] = 'STATISTIQUES GÉNÉRALES'
        ws['A6'].font = Font(size=14, bold=True)
        
        stats = data['statistics']
        row = 8
        
        # Tableau des statistiques
        headers = ['Métrique', 'Valeur']
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=1, value='Total des appareils')
        ws.cell(row=row, column=2, value=stats['total_devices'])
        
        row += 1
        ws.cell(row=row, column=1, value='Appareils en ligne')
        ws.cell(row=row, column=2, value=stats['online_devices'])
        
        row += 1
        ws.cell(row=row, column=1, value='Appareils hors ligne')
        ws.cell(row=row, column=2, value=stats['offline_devices'])
        
        if 'scans_today' in stats:
            row += 1
            ws.cell(row=row, column=1, value='Scans effectués aujourd\'hui')
            ws.cell(row=row, column=2, value=stats['scans_today'])
        
        # Taux de disponibilité
        if stats['total_devices'] > 0:
            row += 1
            uptime_percentage = (stats['online_devices'] / stats['total_devices']) * 100
            ws.cell(row=row, column=1, value='Taux de disponibilité')
            ws.cell(row=row, column=2, value=f"{uptime_percentage:.1f}%")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _add_excel_devices_sheet(self, wb, data):
        """Ajoute la feuille des appareils au rapport Excel"""
        ws = wb.create_sheet("Appareils")
        
        # Titre
        ws['A1'] = 'LISTE DES APPAREILS'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # En-têtes
        headers = ['Adresse IP', 'Nom d\'hôte', 'Statut', 'Dernière vue']
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Données
        row = 4
        for device in data['devices']:
            ws.cell(row=row, column=1, value=device['ip'])
            ws.cell(row=row, column=2, value=device['hostname'])
            ws.cell(row=row, column=3, value=device['status'])
            ws.cell(row=row, column=4, value=device['last_seen'])
            
            # Couleur selon le statut
            if device['status'] == 'online':
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill
            
            row += 1
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
    
    def _add_excel_summary_sheet(self, wb, data):
        """Ajoute la feuille de résumé au rapport Excel"""
        ws = wb.create_sheet("Résumé")
        
        # Titre
        ws['A1'] = 'RÉSUMÉ EXÉCUTIF'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')
        
        # Informations générales
        ws['A3'] = 'Informations générales'
        ws['A3'].font = Font(size=12, bold=True)
        
        row = 5
        ws.cell(row=row, column=1, value='Date du rapport')
        ws.cell(row=row, column=2, value=data['date'])
        row += 1
        
        ws.cell(row=row, column=1, value='Heure de génération')
        ws.cell(row=row, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M"))
        row += 1
        
        stats = data['statistics']
        ws.cell(row=row, column=1, value='Total des équipements surveillés')
        ws.cell(row=row, column=2, value=stats['total_devices'])
        row += 1
        
        ws.cell(row=row, column=1, value='Équipements opérationnels')
        ws.cell(row=row, column=2, value=stats['online_devices'])
        row += 1
        
        ws.cell(row=row, column=1, value='Équipements en panne')
        ws.cell(row=row, column=2, value=stats['offline_devices'])
        row += 1
        
        if stats['total_devices'] > 0:
            uptime_percentage = (stats['online_devices'] / stats['total_devices']) * 100
            ws.cell(row=row, column=1, value='Taux de disponibilité')
            ws.cell(row=row, column=2, value=f"{uptime_percentage:.1f}%")
            
            # Évaluation de la santé du réseau
            row += 2
            ws.cell(row=row, column=1, value='Évaluation de la santé du réseau')
            ws.cell(row=row, column=1).font = Font(size=12, bold=True)
            row += 1
            
            if uptime_percentage >= 95:
                status = "Excellent"
                color = "90EE90"
            elif uptime_percentage >= 85:
                status = "Bon"
                color = "FFFF99"
            elif uptime_percentage >= 70:
                status = "Moyen"
                color = "FFB366"
            else:
                status = "Critique"
                color = "FF6666"
            
            ws.cell(row=row, column=1, value='Statut global')
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def generate_custom_report(self, data, report_type, filename):
        """
        Génère un rapport personnalisé
        
        Args:
            data (dict): Données du rapport
            report_type (str): Type de rapport ('pdf' ou 'excel')
            filename (str): Nom du fichier
        """
        if report_type.lower() == 'pdf':
            self.generate_pdf_report(data, filename)
        elif report_type.lower() == 'excel':
            self.generate_excel_report(data, filename)
        else:
            raise ValueError("Type de rapport non supporté. Utilisez 'pdf' ou 'excel'.")
    
    def list_reports(self):
        """
        Liste tous les rapports disponibles
        
        Returns:
            list: Liste des fichiers de rapports
        """
        try:
            files = os.listdir(self.reports_dir)
            reports = []
            
            for file in files:
                if file.endswith(('.pdf', '.xlsx')):
                    filepath = os.path.join(self.reports_dir, file)
                    file_info = {
                        'filename': file,
                        'path': filepath,
                        'size': os.path.getsize(filepath),
                        'created': datetime.fromtimestamp(os.path.getctime(filepath))
                    }
                    reports.append(file_info)
            
            # Trier par date de création (plus récent en premier)
            reports.sort(key=lambda x: x['created'], reverse=True)
            
            return reports
            
        except Exception as e:
            print(f"❌ Erreur lors de la liste des rapports: {str(e)}")
            return []
    
    def delete_report(self, filename):
        """
        Supprime un rapport
        
        Args:
            filename (str): Nom du fichier à supprimer
            
        Returns:
            bool: True si supprimé avec succès
        """
        try:
            filepath = os.path.join(self.reports_dir, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
                print(f"🗑️ Rapport supprimé: {filename}")
                return True
            else:
                print(f"⚠️ Fichier non trouvé: {filename}")
                return False
                
        except Exception as e:
            print(f"❌ Erreur lors de la suppression: {str(e)}")
            return False
    
    def generate_ai_report(self, ai_data, filename):
        """
        Génère un rapport spécifique aux analyses IA
        
        Args:
            ai_data (dict): Données d'analyse IA
            filename (str): Nom du fichier JSON
        """
        try:
            import json
            
            # Ajouter la date de génération
            ai_data['generated_at'] = datetime.now().isoformat()
            ai_data['report_type'] = 'ai_analysis'
            
            # Sauvegarder en JSON
            filepath = os.path.join(self.reports_dir, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(ai_data, f, indent=2, ensure_ascii=False)
            
            print(f"🧠 Rapport IA généré: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du rapport IA: {str(e)}")
            return None 

    def generate_report(self, report_type='daily', format='pdf', date_from=None, date_to=None, description='', models=None):
        """
        Génère un rapport complet selon le type demandé
        
        Args:
            report_type (str): Type de rapport ('daily', 'weekly', 'monthly', 'custom')
            format (str): Format du rapport ('pdf' ou 'excel')
            date_from (str): Date de début (optionnel)
            date_to (str): Date de fin (optionnel)
            description (str): Description du rapport
            models (dict): Dictionnaire contenant les modèles (Device, ScanHistory, Alert, db)
            
        Returns:
            str: Chemin du fichier généré ou None si erreur
        """
        try:
            # Vérifier que les modèles sont fournis
            if not models:
                print("❌ Erreur: Modèles non fournis pour la génération du rapport")
                return None
            
            Device = models.get('Device')
            ScanHistory = models.get('ScanHistory')
            Alert = models.get('Alert')
            db = models.get('db')
            
            if not all([Device, ScanHistory, Alert, db]):
                print("❌ Erreur: Modèles incomplets pour la génération du rapport")
                return None
            
            # Récupérer les données selon le type de rapport
            if report_type == 'daily':
                # Rapport quotidien
                devices = Device.query.all()
                today = datetime.now().date()
                scans_today = ScanHistory.query.filter(
                    db.func.date(ScanHistory.timestamp) == today
                ).count()
                
            elif report_type == 'weekly':
                # Rapport hebdomadaire
                devices = Device.query.all()
                week_ago = datetime.now().date() - timedelta(days=7)
                scans_this_week = ScanHistory.query.filter(
                    ScanHistory.timestamp >= week_ago
                ).count()
                
            elif report_type == 'monthly':
                # Rapport mensuel
                devices = Device.query.all()
                month_ago = datetime.now().date() - timedelta(days=30)
                scans_this_month = ScanHistory.query.filter(
                    ScanHistory.timestamp >= month_ago
                ).count()
                
            else:
                # Rapport personnalisé
                devices = Device.query.all()
                scans_today = 0
            
            # Préparer les données du rapport
            total_devices = len(devices)
            online_devices = sum(1 for d in devices if d.is_online)
            offline_devices = total_devices - online_devices
            
            # Statistiques
            stats = {
                'total_devices': total_devices,
                'online_devices': online_devices,
                'offline_devices': offline_devices,
                'scans_today': scans_today if report_type == 'daily' else 0
            }
            
            # Données des appareils
            devices_data = []
            for device in devices:
                devices_data.append({
                    'ip': device.ip,
                    'hostname': device.hostname or 'Unknown',
                    'status': 'online' if device.is_online else 'offline',
                    'last_seen': device.last_seen.strftime('%d/%m/%Y %H:%M') if device.last_seen else 'Never',
                    'device_type': device.device_type,
                    'health_score': device.health_score
                })
            
            # Données complètes du rapport
            report_data = {
                'date': datetime.now().strftime('%d/%m/%Y'),
                'report_type': report_type,
                'description': description,
                'statistics': stats,
                'devices': devices_data
            }
            
            # Générer le fichier selon le format
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"rapport_{report_type}_{timestamp}.{format}"
            
            if format.lower() == 'pdf':
                self.generate_pdf_report(report_data, filename)
            elif format.lower() == 'excel':
                self.generate_excel_report(report_data, filename)
            else:
                raise ValueError(f"Format non supporté: {format}")
            
            return os.path.join(self.reports_dir, filename)
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du rapport: {str(e)}")
            return None 